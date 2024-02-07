import os
from openpyxl import Workbook
from main import reading_and_processing
from variables import fileDetail
from util import folder_selection, display_column_names, custom_folder_selection
from step import Isc_20_mA_step_size, Turn_off_80mA_step_size, Turn_off_80mA_HL_step_size, Rf_step_size, Rr_step_size

def main():
    print()
    custom_folder_selection()

    print()
    display_column_names()

    print()
    temp_fast_track = ''
    while (temp_fast_track == '') :
        temp_fast_track = input('Is this a rush / fast-track processing? (y/n): ')
    if temp_fast_track == 'y' or temp_fast_track == 'Y' or temp_fast_track == 'Yes':
        fileDetail['fast_track'] = True
    else:
        fileDetail['fast_track'] = False

    print()
    if (fileDetail['selected_column']['Isc_20mA'] == 1):
        Isc_20_mA_step_size()
        # Create a new file and new folder for storing charts
        fileDetail['isc_fdpath'] = f'{fileDetail["fdpath"]}/Isc_20mA'
        os.mkdir(fileDetail['isc_fdpath'])
        fileDetail['isc_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Isc_20mA_wmap'
        os.mkdir(fileDetail['isc_fdpath_wmap'])
        wb = Workbook()
        wb.save(f'{fileDetail["fdpath"]}/Isc_20mA.xlsx')
        wb_wmap = Workbook()
        wb_wmap.save(f'{fileDetail["fdpath"]}/Isc_20mA_wmap.xlsx')
    if (fileDetail['selected_column']['Turn_off_80mA_'] == 1):
        Turn_off_80mA_step_size()
        # Create a new file and new folder for storing charts
        fileDetail['turn_off_fdpath'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_'
        os.mkdir(fileDetail['turn_off_fdpath'])
        fileDetail['turn_off_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Turn_off_80mA__wmap'
        os.mkdir(fileDetail['turn_off_fdpath_wmap'])
        wb = Workbook()
        wb.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_.xlsx')
        wb_wmap = Workbook()
        wb_wmap.save(f'{fileDetail["fdpath"]}/Turn_off_80mA__wmap.xlsx')
    if (fileDetail['selected_column']['Turn_off_80mA_HL'] == 1):
        Turn_off_80mA_HL_step_size()
        # Create a new file and new folder for storing charts
        fileDetail['turn_off_HL_fdpath'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_HL'
        os.mkdir(fileDetail['turn_off_HL_fdpath'])
        fileDetail['turn_off_HL_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Turn_off_80mA_HL_wmap'
        os.mkdir(fileDetail['turn_off_HL_fdpath_wmap'])
        wb = Workbook()
        wb.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_HL.xlsx')
        wb_wmap = Workbook()
        wb_wmap.save(f'{fileDetail["fdpath"]}/Turn_off_80mA_HL_wmap.xlsx')
    if (fileDetail['selected_column']['Rf'] == 1):
        Rf_step_size()
        # Create a new file and new folder for storing charts
        fileDetail['rf_fdpath'] = f'{fileDetail["fdpath"]}/Rf'
        os.mkdir(fileDetail['rf_fdpath'])
        fileDetail['rf_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Rf_wmap'
        os.mkdir(fileDetail['rf_fdpath_wmap'])
        wb = Workbook()
        wb.save(f'{fileDetail["fdpath"]}/Rf.xlsx')
        wb_wmap = Workbook()
        wb_wmap.save(f'{fileDetail["fdpath"]}/Rf_wmap.xlsx')
    if (fileDetail['selected_column']['Rr'] == 1):
        Rr_step_size()  
        # Create a new file and new folder for storing charts
        fileDetail['rr_fdpath'] = f'{fileDetail["fdpath"]}/Rr'
        os.mkdir(fileDetail['rr_fdpath'])
        fileDetail['rr_fdpath_wmap'] = f'{fileDetail["fdpath"]}/Rr_wmap'
        os.mkdir(fileDetail['rr_fdpath_wmap'])
        wb = Workbook()
        wb.save(f'{fileDetail["fdpath"]}/Rr.xlsx')
        wb_wmap = Workbook()
        wb_wmap.save(f'{fileDetail["fdpath"]}/Rr_wmap.xlsx')
    
    print()
    reading_and_processing()
    print('All done!!!!')

main()